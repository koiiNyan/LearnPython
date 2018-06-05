from openpyxl import load_workbook

def read_excel(filename):
    workbook = load_workbook(filename)
    worksheet = workbook['Sheet0']

    
    excel_data = list()

    for row in range(2,40):
        excel_row = dict()
        if worksheet.cell(row=row, column=2).value is None:
            break

        excel_row['District'] = worksheet.cell(row=row, column=6).value
        excel_row['Location'] = worksheet.cell(row=row, column=7).value
        excel_row['Working_Hours'] = worksheet.cell(row=row, column=12).value

        excel_data.append(excel_row)
        

    return excel_data



def show_excel(data):
    print(data)


if __name__ == '__main__':
    excel_data = read_excel('sobaki.xlsx')
    show_excel(excel_data)
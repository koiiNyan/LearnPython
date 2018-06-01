import csv
from openpyxl import Workbook

def read_csv(filename):
    data = []
    with open('participants.csv', 'r', encoding='utf-8') as f:
        fields = ['Name', 'Visited', 'Is Working', 'Company', 'Experience']
        reader = csv.DictReader(f, fields, delimiter = ';')

        for row in reader:
            data.append(row)

    return data


def excel_write(data):
    workbook = Workbook() #creating a new workbook
    worksheet = workbook.active #active worksheet
    worksheet.title = 'Посещаемость' #naming a worksheet

    worksheet.cell(row=1, column=1).value = 'Имя' #znachenie pervoi yacheiki na liste
    worksheet.cell(row=1, column=2).value = 'Пришел?'
    worksheet.cell(row=1, column=3).value = 'Род деятельности'
    worksheet.cell(row=1, column=4).value = 'Компания'
    worksheet.cell(row=1, column=5).value = 'Опыт питона'


    row = 2

    for item in data:
        worksheet.cell(row=row, column=1).value = item['Name']
        worksheet.cell(row=row, column=2).value = item['Visited']
        worksheet.cell(row=row, column=3).value = item['Is Working']
        worksheet.cell(row=row, column=4).value = item['Company']
        worksheet.cell(row=row, column=5).value = item['Experience']
        row +=1
  



    workbook.save('participants.xlsx')


csv_data = read_csv('participants.csv')
excel_write(csv_data)